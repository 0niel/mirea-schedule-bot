from kutana import Plugin, get_path
from bs4 import BeautifulSoup
from database import schedule_collection
from openpyxl.utils import get_column_letter
from datetime import datetime
from zipfile import BadZipFile
import openpyxl
import numpy as np
import requests
import asyncio
import json
import pytz
import os

plugin = Plugin(name="Parser", description="Parser for schedule bot")


async def new_schedule_check():
    while True:
        try:
            resp = requests.get("https://www.mirea.ru/schedule/")
            soup = BeautifulSoup(resp.text, 'lxml')

            test = soup.findAll('a', class_='uk-link-toggle')
            url = test[15].get('href')

            f = open(get_path(__file__, "schedule.xlsx"), "wb")
            ufr = requests.get(url)
            f.write(ufr.content)
            f.close()

            wb = openpyxl.load_workbook(get_path(__file__, "schedule.xlsx"))
            ws = wb.active

            column_letters = {'schedule': '', 'cabinets': '', 'lesson_type': '', 'teachers': ''}

            for column in range(50, 900):
                column_letter = get_column_letter(column)
                current_column = ws[column_letter]
                for cell in current_column:
                    if isinstance(cell.value, str):
                        if cell.value.find(json.load(open("config.json", 'r'))['group']) > -1:
                            column_letters['schedule'] = column_letter
                            column_letters['cabinets'] = get_column_letter(column + 3)
                            column_letters['lesson_type'] = get_column_letter(column + 1)
                            column_letters['teachers'] = get_column_letter(column + 2)

            columns = {'schedule': ws[column_letters['schedule']], 'cabinets': ws[column_letters['cabinets']],
                       'lesson_type': ws[column_letters['lesson_type']], 'teachers': ws[column_letters['teachers']]}

            column_ws_array = [[], [], [], []]

            for cell in columns['schedule']:
                column_ws_array[0].append(cell.value)

            for cell in columns['cabinets']:
                column_ws_array[1].append(cell.value)

            for cell in columns['lesson_type']:
                column_ws_array[2].append(cell.value)

            for cell in columns['teachers']:
                column_ws_array[3].append(cell.value)

            cabs = np.array_split(column_ws_array[1][3:75], 6)
            schedule = np.array_split(column_ws_array[0][3:75], 6)
            lesson_types = np.array_split(column_ws_array[2][3:75], 6)
            teachers = np.array_split(column_ws_array[3][3:75], 6)

            os.remove(get_path(__file__, "schedule.xlsx"))

            all_days_list, all_cabs_list, all_types_list, all_teachers_list = [], [], [], []

            for day in schedule:
                all_days_list.append(list(day))

            for cab in cabs:
                all_cabs_list.append(list(cab))

            for lesson_type in lesson_types:
                all_types_list.append(list(lesson_type))

            for teacher in teachers:
                all_teachers_list.append(list(teacher))

            schedule_all = schedule_collection.find_one({'type': 'schedule'})

            tz = pytz.timezone('Europe/Moscow')
            date = datetime.now(tz).weekday()

            if schedule_all is not None:
                if date == 0:
                    week_number = schedule_all['week_number'] + 1
                    is_even = True if week_number % 2 == 0 else False
                    schedule_collection.update_one({'type': 'schedule'},
                                                   {'$set': {'is_even': is_even, 'week_number': week_number}})

                schedule_collection.update_one({'type': 'schedule'}, {'$set': {'schedule': all_days_list,
                                                                               'cabinets': all_cabs_list,
                                                                               'lesson_types': all_types_list,
                                                                               'teachers': all_teachers_list}})
            else:
                schedule_collection.insert_one({'type': 'schedule',
                                                'schedule': all_days_list,
                                                'cabinets': all_cabs_list,
                                                'lesson_types': all_types_list,
                                                'teachers': all_teachers_list,
                                                'is_even': False, 'week_number': 1})
        except BadZipFile:
            print('Exception BadZipFile: file not found on mirea.ru')

        await asyncio.sleep(600)


@plugin.on_start()
async def _(app):
    backend = app.get_backends()[0]

    # Run only if first backend is Vkontakte
    if backend.get_identity() == "vkontakte":
        asyncio.ensure_future(new_schedule_check())
